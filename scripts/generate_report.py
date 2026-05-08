from __future__ import annotations

from collections import Counter
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils import FIELDS, LATEST_CSV, WEEKLY_MD, WEEKLY_XLSX, ensure_dirs, read_csv, today_str


PRIORITY_ORDER = {
    "P0": 0,
    "P1": 1,
    "P2": 2,
    "P3": 3,
}


def _to_int(value: str) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _priority_rank(row: Dict[str, str]) -> int:
    prefix = row.get("推荐参加优先级", "P9")[:2]
    return PRIORITY_ORDER.get(prefix, 9)


def sort_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    return sorted(
        rows,
        key=lambda row: (
            _priority_rank(row),
            -_to_int(row.get("金融路线价值评分", "0")),
            -_to_int(row.get("数据分析量化评分", "0")),
            -_to_int(row.get("含金量分", "0")),
            row.get("比赛名称", ""),
        ),
    )


def _escape_md(value: str) -> str:
    return str(value).replace("|", "\\|")


def generate_markdown(rows: List[Dict[str, str]]) -> str:
    sorted_rows = sort_rows(rows)
    category_counts = Counter(row.get("类别", "未分类") or "未分类" for row in sorted_rows)
    priority_counts = Counter(row.get("推荐参加优先级", "未评级") or "未评级" for row in sorted_rows)
    top_rows = sorted_rows[:12]

    lines = [
        "# 大学生竞赛情报周报",
        "",
        f"- 生成日期：{today_str()}",
        f"- 收录竞赛：{len(sorted_rows)} 个",
        f"- P0/P1 推荐：{sum(1 for row in sorted_rows if row.get('推荐参加优先级', '').startswith(('P0', 'P1')))} 个",
        "",
        "## 本周优先关注",
        "",
        "| 优先级 | 比赛名称 | 类别 | 含金量 | 金融 | 数据/量化 | 数模 | 官方链接 |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]

    for row in top_rows:
        link = row.get("官方链接", "")
        link_text = f"[查看]({link})" if link else ""
        lines.append(
            "| {priority} | {name} | {category} | {prestige} | {finance} | {data} | {modeling} | {link} |".format(
                priority=_escape_md(row.get("推荐参加优先级", "")),
                name=_escape_md(row.get("比赛名称", "")),
                category=_escape_md(row.get("类别", "")),
                prestige=_escape_md(f"{row.get('含金量评级', '')}/{row.get('含金量分', '')}"),
                finance=_escape_md(row.get("金融路线价值评分", "")),
                data=_escape_md(row.get("数据分析量化评分", "")),
                modeling=_escape_md(row.get("数学建模能力评分", "")),
                link=link_text,
            )
        )

    lines.extend(["", "## 类别分布", ""])
    for category, count in category_counts.most_common():
        lines.append(f"- {category}：{count} 个")

    lines.extend(["", "## 推荐优先级分布", ""])
    for priority, count in sorted(priority_counts.items()):
        lines.append(f"- {priority}：{count} 个")

    lines.extend(["", "## 完整清单", ""])
    for category in category_counts:
        lines.extend([f"### {category}", ""])
        for row in [item for item in sorted_rows if item.get("类别", "未分类") == category]:
            lines.append(
                "- **{name}**：{priority}，金融 {finance}，商赛 {business}，数模 {modeling}，数据/量化 {data}。{reason}".format(
                    name=row.get("比赛名称", ""),
                    priority=row.get("推荐参加优先级", ""),
                    finance=row.get("金融路线价值评分", ""),
                    business=row.get("商赛价值评分", ""),
                    modeling=row.get("数学建模能力评分", ""),
                    data=row.get("数据分析量化评分", ""),
                    reason=row.get("推荐理由", ""),
                )
            )
        lines.append("")

    return "\n".join(lines).strip() + "\n"


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 10), 38)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def generate_excel(rows: List[Dict[str, str]]) -> None:
    sorted_rows = sort_rows(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "本周推荐"
    focus_fields = [
        "推荐参加优先级",
        "比赛名称",
        "类别",
        "含金量评级",
        "含金量分",
        "金融路线价值评分",
        "商赛价值评分",
        "数学建模能力评分",
        "数据分析量化评分",
        "报名时间",
        "比赛时间",
        "官方链接",
        "推荐理由",
    ]
    ws.append(focus_fields)
    for row in sorted_rows:
        ws.append([row.get(field, "") for field in focus_fields])
    _style_sheet(ws)

    full = wb.create_sheet("完整数据")
    full.append(FIELDS)
    for row in sorted_rows:
        full.append([row.get(field, "") for field in FIELDS])
    _style_sheet(full)

    guide = wb.create_sheet("评分说明")
    guide.append(["维度", "说明"])
    guide.append(["含金量", "结合主办单位、全国/国际属性、权威学会或教育部门背景。"])
    guide.append(["金融路线", "关注金融、量化、风控、投研、证券、银行、CFA、财务分析等关键词。"])
    guide.append(["商赛简历", "关注商业案例、市场调查、咨询、创业、管理、营销等关键词。"])
    guide.append(["数学建模", "关注数学建模、统计建模、运筹优化、预测、仿真和论文表达。"])
    guide.append(["数据/量化", "关注数据分析、统计、算法、AI、机器学习、Python、量化研究等能力。"])
    _style_sheet(guide)

    wb.save(WEEKLY_XLSX)


def main() -> None:
    ensure_dirs()
    rows = read_csv(LATEST_CSV)
    markdown = generate_markdown(rows)
    WEEKLY_MD.write_text(markdown, encoding="utf-8")
    generate_excel(rows)
    print(f"Generated report for {len(rows)} competitions.")


if __name__ == "__main__":
    main()

