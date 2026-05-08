from __future__ import annotations

import datetime as dt
import re
from collections import Counter
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils import (
    LATEST_CSV,
    SCHEDULE_MD,
    SCHEDULE_XLSX,
    ensure_dirs,
    normalize_space,
    now_beijing_str,
    read_csv,
    today_beijing_str,
)


EVENT_FIELDS = [
    "事件类型",
    "解析日期",
    "时间状态",
    "时间原文",
    "比赛名称",
    "类别",
    "推荐参加优先级",
    "含金量评级",
    "金融路线价值评分",
    "商赛价值评分",
    "数学建模能力评分",
    "数据分析量化评分",
    "官方链接",
    "来源页面",
    "是否疑似误抓取",
    "备注",
]

DATE_FULL_PATTERNS = [
    re.compile(r"(?P<year>20\d{2})\s*年\s*(?P<month>\d{1,2})\s*月\s*(?P<day>\d{1,2})\s*日?"),
    re.compile(r"(?P<year>20\d{2})[./-](?P<month>\d{1,2})[./-](?P<day>\d{1,2})"),
]
DATE_MONTH_PATTERNS = [
    re.compile(r"(?P<year>20\d{2})\s*年\s*(?P<month>\d{1,2})\s*月"),
    re.compile(r"(?P<year>20\d{2})[./-](?P<month>\d{1,2})(?![./-]\d)"),
]
DATE_YEAR_PATTERNS = [
    re.compile(r"(?P<year>20\d{2})\s*年"),
]

RANGE_CONNECTOR_RE = re.compile(r"\s*(?:至|到|—|－|-|~|～)\s*")
UNCERTAIN_PATTERNS = [
    "通常每年",
    "每年上半年",
    "每年下半年",
    "每年春季",
    "每年秋季",
    "每年",
    "以年度通知为准",
    "赛后数月公示",
    "关注学校通知",
    "待核实",
    "待核验",
    "待定",
    "另行通知",
    "暂未公布",
]

INVALID_NAME_PATTERNS = [
    "中国知网",
    "SPSSPRO",
    "CUMCM通讯",
    "《数学建模及其应用》",
    "数学建模及其应用",
    "中国高校数学建模课程中心",
    "竞赛组织",
    "苏州同元软控信息技术有限公司",
    "大赛通知",
    "挑战杯动态",
    "竞赛章程",
    "竞赛简介",
    "培训交流",
    "学霸笔记",
    "官方网站 移动版",
    "Competitions - DataFountain",
    "数据科学竞赛/大数据 Competitions - DataFountain",
    "数据科学竞赛/大数据",
]
NAVIGATION_OR_NEWS_PATTERNS = [
    "动态",
    "通知",
    "资讯",
    "简介",
    "章程",
    "培训",
    "倒计时",
    "笔记",
    "课程中心",
    "期刊",
    "通讯",
    "合作",
    "平台",
    "栏目",
    "频道",
]
GENERIC_TOPIC_TITLES = {"大数据和人工智能", "数据科学竞赛/大数据 Competitions - DataFountain"}
VALID_COMPETITION_KEYWORDS = ["竞赛", "大赛", "挑战赛", "Challenge", "Cup", "Contest", "赛"]
GENERIC_NAVIGATION_PATTERNS = ["首页", "通知公告", "新闻动态", "下载中心", "关于我们", "联系我们"]
MOJIBAKE_RE = re.compile(r"[�ÃÂåäæçÐð]{2,}|\\x[0-9a-fA-F]{2}")
MOJIBAKE_CHARS = set("�ÃÂåäæèç")


def _to_int(value: str) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _priority_rank(value: str) -> int:
    if value.startswith("P0"):
        return 0
    if value.startswith("P1"):
        return 1
    if value.startswith("P2"):
        return 2
    if value.startswith("P3"):
        return 3
    return 9


def _sort_key(event: Dict[str, str]) -> Tuple[str, int, int, str]:
    parsed_date = event.get("解析日期") or "9999-12-31"
    return (
        parsed_date,
        _priority_rank(event.get("推荐参加优先级", "")),
        -_to_int(event.get("金融路线价值评分", "0")),
        event.get("比赛名称", ""),
    )


def _escape_md(value: str) -> str:
    return str(value or "").replace("|", "\\|")


def _md_link(url: str) -> str:
    return f"[查看]({url})" if url else ""


def _valid_date(year: int, month: int, day: int) -> str:
    try:
        return dt.date(year, month, day).isoformat()
    except ValueError:
        return ""


def _candidate_from_match(match: re.Match[str], precision: str, priority: int) -> Dict[str, object]:
    year = int(match.group("year"))
    month = int(match.groupdict().get("month") or 1)
    day = int(match.groupdict().get("day") or 1)
    parsed = _valid_date(year, month, day)
    return {
        "start": match.start(),
        "end": match.end(),
        "precision": precision,
        "priority": priority,
        "parsed_date": parsed,
    }


def _date_candidates(text: str) -> List[Dict[str, object]]:
    candidates: List[Dict[str, object]] = []
    for pattern in DATE_FULL_PATTERNS:
        candidates.extend(_candidate_from_match(match, "具体日期", 0) for match in pattern.finditer(text))
    for pattern in DATE_MONTH_PATTERNS:
        candidates.extend(_candidate_from_match(match, "仅到月份", 1) for match in pattern.finditer(text))
    for pattern in DATE_YEAR_PATTERNS:
        candidates.extend(_candidate_from_match(match, "仅到年份", 2) for match in pattern.finditer(text))
    candidates.sort(key=lambda item: (int(item["start"]), int(item["priority"]), -int(item["end"])))

    selected: List[Dict[str, object]] = []
    occupied: List[Tuple[int, int]] = []
    for candidate in candidates:
        start = int(candidate["start"])
        end = int(candidate["end"])
        if any(not (end <= taken_start or start >= taken_end) for taken_start, taken_end in occupied):
            continue
        selected.append(candidate)
        occupied.append((start, end))
    selected.sort(key=lambda item: int(item["start"]))
    return selected


def _looks_like_invalid_date(text: str) -> bool:
    return bool(re.search(r"20\d{2}\s*(?:年|[./-])\s*\d{1,2}", text))


def parse_time_text(value: str) -> Dict[str, str]:
    raw = normalize_space(value)
    if not raw:
        return {"解析日期": "", "时间状态": "周期性/待核实", "备注": "时间字段为空。"}
    if any(pattern in raw for pattern in UNCERTAIN_PATTERNS):
        return {"解析日期": "", "时间状态": "周期性/待核实", "备注": "时间为周期性描述或需回源确认。"}

    candidates = _date_candidates(raw)
    if not candidates:
        status = "疑似误提取" if _looks_like_invalid_date(raw) else "周期性/待核实"
        note = "检测到疑似日期但无法解析。" if status == "疑似误提取" else "未识别到可排序日期。"
        return {"解析日期": "", "时间状态": status, "备注": note}

    first = candidates[0]
    parsed_date = str(first["parsed_date"])
    if not parsed_date:
        return {"解析日期": "", "时间状态": "疑似误提取", "备注": "日期数值超出有效范围。"}

    status = str(first["precision"])
    if len(candidates) >= 2:
        gap = raw[int(first["end"]) : int(candidates[1]["start"])]
        if RANGE_CONNECTOR_RE.fullmatch(gap):
            status = "日期范围"
    else:
        tail = raw[int(first["end"]) : int(first["end"]) + 12]
        if RANGE_CONNECTOR_RE.match(tail):
            status = "日期范围"

    note = ""
    if status in {"仅到月份", "仅到年份"}:
        note = "解析日期仅用于排序，不建议直接导入日历。"
    return {"解析日期": parsed_date, "时间状态": status, "备注": note}


def invalid_competition_reason(row: Dict[str, str]) -> str:
    name = normalize_space(row.get("比赛名称", ""))
    link = normalize_space(row.get("官方链接", ""))
    if not name:
        return "比赛名称为空。"
    if MOJIBAKE_RE.search(name) or sum(1 for char in name if char in MOJIBAKE_CHARS) >= 3:
        return "标题疑似编码异常或乱码。"
    if name in GENERIC_TOPIC_TITLES:
        return "标题是平台栏目或泛化主题，不是明确比赛名称。"
    for pattern in NAVIGATION_OR_NEWS_PATTERNS:
        if pattern.lower() in name.lower():
            return f"标题疑似导航/资讯/培训内容：{pattern}。"
    for pattern in INVALID_NAME_PATTERNS:
        if pattern.lower() in name.lower():
            return f"标题命中误抓取关键词：{pattern}。"
    if name in GENERIC_NAVIGATION_PATTERNS:
        return "标题疑似官网导航。"
    if any(keyword.lower() in name.lower() for keyword in VALID_COMPETITION_KEYWORDS):
        return ""
    if "datafountain.cn/competitions/" in link and re.search(r"/competitions/\d+", link):
        return ""
    return "标题缺少竞赛/大赛/挑战赛/Challenge/Cup/Contest/赛等比赛关键词。"


def is_valid_competition(row: Dict[str, str]) -> bool:
    return invalid_competition_reason(row) == ""


def _suspect_entry(row: Dict[str, str], reason: str) -> Dict[str, str]:
    time_parts = []
    for label in ["报名时间", "比赛时间", "结果公布时间"]:
        value = normalize_space(row.get(label, ""))
        if value:
            time_parts.append(f"{label}: {value}")
    return {
        "事件类型": "疑似误抓取",
        "解析日期": "",
        "时间状态": "疑似误抓取",
        "时间原文": "；".join(time_parts),
        "比赛名称": normalize_space(row.get("比赛名称", "")),
        "类别": normalize_space(row.get("类别", "")),
        "推荐参加优先级": normalize_space(row.get("推荐参加优先级", "")),
        "含金量评级": normalize_space(row.get("含金量评级", "")),
        "金融路线价值评分": normalize_space(row.get("金融路线价值评分", "")),
        "商赛价值评分": normalize_space(row.get("商赛价值评分", "")),
        "数学建模能力评分": normalize_space(row.get("数学建模能力评分", "")),
        "数据分析量化评分": normalize_space(row.get("数据分析量化评分", "")),
        "官方链接": normalize_space(row.get("官方链接", "")),
        "来源页面": normalize_space(row.get("来源页面", "")),
        "是否疑似误抓取": "是",
        "备注": reason,
    }


def build_events(rows: Iterable[Dict[str, str]]) -> Tuple[List[Dict[str, str]], Dict[str, int], List[Dict[str, str]]]:
    event_specs = [
        ("报名/报名截止", "报名时间"),
        ("正式比赛", "比赛时间"),
        ("结果公布/获奖公示", "结果公布时间"),
    ]
    events: List[Dict[str, str]] = []
    stats = {"空时间字段跳过": 0}
    suspect_entries: List[Dict[str, str]] = []
    for row in rows:
        invalid_reason = invalid_competition_reason(row)
        if invalid_reason:
            suspect_entries.append(_suspect_entry(row, invalid_reason))
        for event_type, time_field in event_specs:
            raw_time = normalize_space(row.get(time_field, ""))
            if not raw_time:
                stats["空时间字段跳过"] += 1
                continue
            parsed = parse_time_text(raw_time)
            note_parts = [invalid_reason] if invalid_reason else [parsed["备注"]]
            events.append(
                {
                    "事件类型": event_type,
                    "解析日期": parsed["解析日期"],
                    "时间状态": parsed["时间状态"],
                    "时间原文": raw_time,
                    "比赛名称": normalize_space(row.get("比赛名称", "")),
                    "类别": normalize_space(row.get("类别", "")),
                    "推荐参加优先级": normalize_space(row.get("推荐参加优先级", "")),
                    "含金量评级": normalize_space(row.get("含金量评级", "")),
                    "金融路线价值评分": normalize_space(row.get("金融路线价值评分", "")),
                    "商赛价值评分": normalize_space(row.get("商赛价值评分", "")),
                    "数学建模能力评分": normalize_space(row.get("数学建模能力评分", "")),
                    "数据分析量化评分": normalize_space(row.get("数据分析量化评分", "")),
                    "官方链接": normalize_space(row.get("官方链接", "")),
                    "来源页面": normalize_space(row.get("来源页面", "")),
                    "是否疑似误抓取": "是" if invalid_reason else "否",
                    "备注": " ".join(note_parts),
                }
            )
    return events, stats, suspect_entries


def _valid_events(events: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    return [event for event in events if event["是否疑似误抓取"] == "否"]


def _suspect_events(events: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    return [event for event in events if event["是否疑似误抓取"] == "是"]


def _scheduled_events(events: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    return sorted(
        [
            event
            for event in events
            if event["解析日期"]
            and event["时间状态"] not in {"周期性/待核实", "疑似误提取"}
        ],
        key=_sort_key,
    )


def _future_events(events: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    today = today_beijing_str()
    return [event for event in events if event.get("解析日期", "") >= today]


def _calendar_events(events: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    return sorted(
        [
            event
            for event in events
            if event["是否疑似误抓取"] == "否"
            and event["比赛名称"]
            and event["事件类型"]
            and event["解析日期"]
            and event["解析日期"] >= today_beijing_str()
            and event["时间状态"] in {"具体日期", "日期范围"}
        ],
        key=_sort_key,
    )


def _uncertain_events(events: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    return sorted(
        [
            event
            for event in events
            if event["是否疑似误抓取"] == "否"
            and (
                event["时间状态"] in {"周期性/待核实", "疑似误提取", "仅到月份", "仅到年份"}
                or not event["解析日期"]
            )
        ],
        key=_sort_key,
    )


def _unique_by_competition(events: Iterable[Dict[str, str]], limit: int = 12) -> List[Dict[str, str]]:
    seen: set[str] = set()
    unique: List[Dict[str, str]] = []
    for event in events:
        name = event["比赛名称"]
        if not name or name in seen:
            continue
        seen.add(name)
        unique.append(event)
        if len(unique) >= limit:
            break
    return unique


def _append_table(lines: List[str], headers: List[str], rows: List[List[str]]) -> None:
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    if not rows:
        lines.append("| " + " | ".join(["暂无"] + [""] * (len(headers) - 1)) + " |")
        return
    for row in rows:
        lines.append("| " + " | ".join(_escape_md(value) for value in row) + " |")


def generate_markdown(events: List[Dict[str, str]], stats: Dict[str, int], suspect_entries: List[Dict[str, str]]) -> str:
    valid_events = _valid_events(events)
    scheduled_valid = _scheduled_events(valid_events)
    future_scheduled_valid = _future_events(scheduled_valid)
    competition_events = [event for event in future_scheduled_valid if event["事件类型"] == "正式比赛"]
    signup_events = [event for event in future_scheduled_valid if event["事件类型"] == "报名/报名截止"]
    result_events = [event for event in future_scheduled_valid if event["事件类型"] == "结果公布/获奖公示"]
    calendar_events = _calendar_events(events)
    uncertain_events = _uncertain_events(events)
    clear_events = [event for event in scheduled_valid if event["时间状态"] in {"具体日期", "日期范围"}]

    lines = [
        "# 大学生竞赛日程报告",
        "",
        f"- 生成时间：北京时间 {now_beijing_str()}",
        "- 自动运行时间：北京时间每周日 00:00",
        "- GitHub Actions cron：0 16 * * 6",
        "- 说明：GitHub Actions 使用 UTC 时间，0 16 * * 6 对应北京时间周日 00:00。",
        "",
        "## 数据质量摘要",
        "",
        f"- 有明确日期的事件：{len(clear_events)} 条",
        f"- 可进日历事件：{len(calendar_events)} 条",
        f"- 待核实事件：{len(uncertain_events)} 条",
        f"- 疑似误抓取事件：{len(suspect_entries)} 条",
        f"- 空时间字段跳过：{stats.get('空时间字段跳过', 0)} 条",
        "",
        "## 一、未来重点比赛时间",
        "",
    ]
    _append_table(
        lines,
        ["日期", "时间状态", "比赛名称", "类别", "优先级", "含金量", "金融", "商赛", "数模", "数据/量化", "官方链接"],
        [
            [
                event["解析日期"],
                event["时间状态"],
                event["比赛名称"],
                event["类别"],
                event["推荐参加优先级"],
                event["含金量评级"],
                event["金融路线价值评分"],
                event["商赛价值评分"],
                event["数学建模能力评分"],
                event["数据分析量化评分"],
                _md_link(event["官方链接"]),
            ]
            for event in competition_events
        ],
    )

    lines.extend(["", "## 二、报名/截止提醒", ""])
    _append_table(
        lines,
        ["日期", "时间状态", "比赛名称", "类别", "优先级", "报名时间原文", "官方链接"],
        [
            [
                event["解析日期"],
                event["时间状态"],
                event["比赛名称"],
                event["类别"],
                event["推荐参加优先级"],
                event["时间原文"],
                _md_link(event["官方链接"]),
            ]
            for event in signup_events
        ],
    )

    lines.extend(["", "## 三、结果公布/获奖公示提醒", ""])
    _append_table(
        lines,
        ["日期", "时间状态", "比赛名称", "类别", "优先级", "结果时间原文", "官方链接"],
        [
            [
                event["解析日期"],
                event["时间状态"],
                event["比赛名称"],
                event["类别"],
                event["推荐参加优先级"],
                event["时间原文"],
                _md_link(event["官方链接"]),
            ]
            for event in result_events
        ],
    )

    lines.extend(["", "## 四、可进日历提醒", ""])
    _append_table(
        lines,
        ["日期", "事件类型", "比赛名称", "备注", "官方链接"],
        [
            [event["解析日期"], event["事件类型"], event["比赛名称"], event["备注"], _md_link(event["官方链接"])]
            for event in calendar_events
        ],
    )

    lines.extend(["", "## 五、待核实时间", ""])
    _append_table(
        lines,
        ["事件类型", "时间状态", "时间原文", "比赛名称", "类别", "优先级", "备注"],
        [
            [
                event["事件类型"],
                event["时间状态"],
                event["时间原文"],
                event["比赛名称"],
                event["类别"],
                event["推荐参加优先级"],
                event["备注"],
            ]
            for event in uncertain_events[:80]
        ],
    )

    lines.extend(["", "## 六、疑似误抓取条目", ""])
    _append_table(
        lines,
        ["比赛名称", "类别", "来源页面", "过滤原因"],
        [
            [event["比赛名称"], event["类别"], event["来源页面"], event["备注"]]
            for event in _unique_by_competition(suspect_entries, limit=80)
        ],
    )

    lines.extend(["", "## 七、本周日程结论", ""])
    top_competitions = _unique_by_competition(competition_events, limit=3)
    top_names = "、".join(event["比赛名称"] for event in top_competitions) or "暂无日期明确的正式比赛"
    needs_check = _unique_by_competition(uncertain_events, limit=5)
    check_names = "、".join(event["比赛名称"] for event in needs_check) or "暂无"
    status_counts = Counter(event["时间状态"] for event in uncertain_events)
    lines.append(f"1. 近期最值得关注的比赛：{top_names}。")
    lines.append(f"2. 需要回源确认时间的比赛：{check_names}。")
    lines.append(
        "3. 当前不适合进日历提醒的比赛事件："
        f"{len(uncertain_events) + len(suspect_entries)} 条，其中待核实 {status_counts.get('周期性/待核实', 0)} 条，"
        f"疑似误抓取 {len(suspect_entries)} 条，空时间字段跳过 {stats.get('空时间字段跳过', 0)} 条。"
    )

    return "\n".join(lines).strip() + "\n"


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 10), 42)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _append_sheet(wb: Workbook, title: str, rows: List[Dict[str, str]], fields: List[str] = EVENT_FIELDS) -> None:
    ws = wb.active if wb.active.title == "Sheet" else wb.create_sheet(title)
    ws.title = title
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    _style_sheet(ws)


def generate_excel(events: List[Dict[str, str]], suspect_entries: List[Dict[str, str]]) -> None:
    valid_events = _valid_events(events)
    scheduled_valid = _future_events(_scheduled_events(valid_events))
    wb = Workbook()
    _append_sheet(wb, "日程总览", sorted(events, key=_sort_key))
    _append_sheet(wb, "比赛时间优先", [event for event in scheduled_valid if event["事件类型"] == "正式比赛"])
    _append_sheet(wb, "报名截止提醒", [event for event in scheduled_valid if event["事件类型"] == "报名/报名截止"])
    _append_sheet(wb, "结果公布提醒", [event for event in scheduled_valid if event["事件类型"] == "结果公布/获奖公示"])
    _append_sheet(wb, "可进日历", _calendar_events(events))
    _append_sheet(wb, "待核实", _uncertain_events(events))
    _append_sheet(wb, "疑似误抓取", sorted(suspect_entries, key=_sort_key))
    wb.save(SCHEDULE_XLSX)


def main() -> None:
    ensure_dirs()
    rows = read_csv(LATEST_CSV)
    events, stats, suspect_entries = build_events(rows)
    SCHEDULE_MD.write_text(generate_markdown(events, stats, suspect_entries), encoding="utf-8")
    generate_excel(events, suspect_entries)
    print(
        "Generated schedule report for "
        f"{len(rows)} competitions, {len(events)} events, and {stats.get('空时间字段跳过', 0)} skipped empty fields."
    )


if __name__ == "__main__":
    main()
